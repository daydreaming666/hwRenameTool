import json
import os.path
import random
import time
import win32api

from rcc.MainWindow import Ui_MainWindow
import rcc.AboutDialog
from PyQt6.QtWidgets import (QMainWindow, QFileDialog, QApplication, QHeaderView, QTableWidgetItem, QMessageBox,
                             QDialog)
from PyQt6.QtCore import (pyqtSlot, Qt, QObject, QThread, pyqtSignal)
from openpyxl import Workbook, load_workbook


class AboutDialog(QDialog, rcc.AboutDialog.Ui_Dialog):
    def __init__(self, parent=None):
        super(AboutDialog, self).__init__(parent)
        self.setupUi(self)


class MainWindow(QMainWindow, Ui_MainWindow):
    import_file_signal = pyqtSignal(str)
    export_config_signal = pyqtSignal(str, dict)
    import_config_signal = pyqtSignal(str)
    rename_signal = pyqtSignal(dict)
    scan_signal = pyqtSignal(str, list, str)

    def __init__(self):
        super(MainWindow, self).__init__()
        self.setupUi(self)

        self.working_directory = os.path.abspath(".")
        self.lineEdit.setText(self.working_directory)
        self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.tableWidget.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)

        self.tableWidget_2.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.tableWidget_2.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        # set working directory
        self.pushButton.clicked.connect(self.set_working_directory)
        # import file
        self.pushButton_2.clicked.connect(self.select_file_import)

        self.worker = Worker()
        self.workThread = QThread()

        self.import_file_signal.connect(self.worker.import_file)
        self.import_config_signal.connect(self.worker.import_config)
        self.export_config_signal.connect(self.worker.export_config)
        self.scan_signal.connect(self.worker.scan_file)
        self.rename_signal.connect(self.worker.rename_file)

        # generate excel file
        self.pushButton_3.clicked.connect(self.worker.generate_excel)
        # import config
        self.pushButton_4.clicked.connect(self.import_config)
        # export config
        self.pushButton_5.clicked.connect(self.export_config)
        # help
        self.pushButton_6.clicked.connect(self.open_help_page)
        # about
        self.pushButton_7.clicked.connect(self.open_about_dialog)
        # scan
        self.pushButton_8.clicked.connect(self.scan_file)
        # rename
        self.pushButton_9.clicked.connect(self.rename_file)

        self.worker.finished_import.connect(self.show_data)
        self.worker.show_error_message.connect(self.show_error_message)
        self.worker.show_message.connect(self.show_message)
        self.worker.set_status_message.connect(self.set_status_message)
        self.worker.update_progress.connect(self.update_progress)

        self.worker.finished_import_config.connect(self.handle_config)
        self.worker.finished_scan.connect(self.handle_scan)

        self.worker.moveToThread(self.workThread)
        self.workThread.start()

    @pyqtSlot()
    def open_about_dialog(self):
        dialog = AboutDialog()
        dialog.exec()

    @pyqtSlot()
    def open_help_page(self):
        # open help page using default browser
        win32api.ShellExecute(0, 'open', "https://github.com/daydreaming666/hwRenameTool",
                              '', '', 1)

    @pyqtSlot()
    def set_working_directory(self):
        directory = QFileDialog.getExistingDirectory(self, "选择工作目录", self.working_directory)
        if directory:
            self.working_directory = directory
            self.lineEdit.setText(directory)

    @pyqtSlot()
    def select_file_import(self):
        file_name = QFileDialog.getOpenFileName(self, "导入文件",
                                                "./exported.xlsx", "Excel Files (*.xlsx)")
        if file_name[0]:
            self.import_file_signal.emit(file_name[0])

    @pyqtSlot(list)
    def show_data(self, data):
        self.tableWidget.setRowCount(len(data))
        for i, row in enumerate(data):
            for j, col in enumerate(row):
                self.tableWidget.setItem(i, j, QTableWidgetItem(str(col)))

    @pyqtSlot(str)
    def show_error_message(self, message):
        QMessageBox.warning(self, "错误", message)

    @pyqtSlot(str)
    def show_message(self, message):
        QMessageBox.information(self, "提示", message)

    @pyqtSlot()
    def export_config(self):
        file_name = QFileDialog.getSaveFileName(self, "导出配置",
                                                "./config.json", "JSON Files (*.json)")
        if file_name[0]:
            config = {"working_directory": self.working_directory,
                      "rename_format": self.lineEdit_2.text(),
                      "data": []}
            for i in range(self.tableWidget.rowCount()):
                config["data"].append([])
                for j in range(self.tableWidget.columnCount()):
                    config["data"][i].append(self.tableWidget.item(i, j).text())
            self.export_config_signal.emit(file_name[0], config)

    @pyqtSlot()
    def import_config(self):
        file_name = QFileDialog.getOpenFileName(self, "导入配置",
                                                "./config.json", "JSON Files (*.json)")
        if file_name[0]:
            self.import_config_signal.emit(file_name[0])

    @pyqtSlot(dict)
    def handle_config(self, config):
        self.working_directory = config["working_directory"]
        self.lineEdit.setText(self.working_directory)
        self.lineEdit_2.setText(config["rename_format"])
        self.tableWidget.setRowCount(len(config["data"]))
        for i, row in enumerate(config["data"]):
            for j, col in enumerate(row):
                self.tableWidget.setItem(i, j, QTableWidgetItem(str(col)))
        self.show_message("配置导入成功")

    @pyqtSlot()
    def rename_file(self):
        if self.tableWidget_2.rowCount() == 0:
            QMessageBox.warning(self, "错误", "请先扫描")
            return
        data = {"working_directory": self.working_directory}
        rename_list = []
        for i in range(self.tableWidget_2.rowCount()):
            old_name = self.tableWidget_2.item(i, 1).text()
            new_name = self.tableWidget_2.item(i, 2).text()
            status = self.tableWidget_2.item(i, 3).text()
            if old_name != new_name and status != "未找到":
                rename_list.append((i, old_name, new_name))
        data["rename_list"] = rename_list
        self.rename_signal.emit(data)

    @pyqtSlot(list)
    def handle_scan(self, data):
        self.tableWidget_2.setRowCount(len(data))
        count = [0, 0]
        for i, row in enumerate(data):
            target_name = row[0]
            old_name = row[1]
            new_name = row[2]
            status = "相同"
            target_name_item = QTableWidgetItem(target_name)
            target_name_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            old_name_item = QTableWidgetItem(old_name)
            old_name_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            new_name_item = QTableWidgetItem(new_name)
            new_name_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            status_item = QTableWidgetItem(status)
            status_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if old_name == new_name == "":
                target_name_item.setForeground(Qt.GlobalColor.red)
                old_name_item.setText("未找到")
                old_name_item.setForeground(Qt.GlobalColor.red)
                new_name_item.setText("不可用")
                new_name_item.setForeground(Qt.GlobalColor.red)
                status_item.setText("未找到")
                status_item.setForeground(Qt.GlobalColor.red)
                count[0] += 1
            elif old_name != new_name:
                target_name_item.setForeground(Qt.GlobalColor.darkBlue)
                old_name_item.setForeground(Qt.GlobalColor.red)
                new_name_item.setForeground(Qt.GlobalColor.darkBlue)
                status_item.setText("须重命名")
                status_item.setForeground(Qt.GlobalColor.red)
                count[1] += 1

            self.tableWidget_2.setItem(i, 0, target_name_item)
            self.tableWidget_2.setItem(i, 1, old_name_item)
            self.tableWidget_2.setItem(i, 2, new_name_item)
            self.tableWidget_2.setItem(i, 3, status_item)
        self.set_status_message("扫描{}项目，{}未找到，{}须重命名".format(
            len(data), count[0], count[1]))
        self.show_message("扫描完成")

    @pyqtSlot(str)
    def set_status_message(self, message):
        self.label_2.setText(message)

    @pyqtSlot()
    def scan_file(self):
        name_list = []
        for i in range(self.tableWidget.rowCount()):
            name_list.append([])
            for j in range(self.tableWidget.columnCount()):
                name_list[i].append(self.tableWidget.item(i, j).text())
        rename_format = self.lineEdit_2.text()
        self.scan_signal.emit(self.working_directory, name_list, rename_format)

    @pyqtSlot(tuple)
    def update_progress(self, status):
        row = status[0]
        update_value = status[1]
        update_status = status[2]
        update_value_item = QTableWidgetItem(str(update_value[0]))
        if update_value[1] == "red":
            color = Qt.GlobalColor.red
        elif update_value[1] == "blue":
            color = Qt.GlobalColor.darkBlue
        else:
            color = Qt.GlobalColor.black
        update_value_item.setForeground(color)
        update_value_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

        update_status_item = QTableWidgetItem(update_status[0])
        if update_status[1] == "red":
            color2 = Qt.GlobalColor.red
        elif update_status[1] == "blue":
            color2 = Qt.GlobalColor.darkBlue
        else:
            color2 = Qt.GlobalColor.black
        update_status_item.setForeground(color2)
        update_status_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

        self.tableWidget_2.setItem(row, 1, update_value_item)
        self.tableWidget_2.setItem(row, 3, update_status_item)
        self.tableWidget_2.scrollToItem(update_value_item)


class Worker(QObject):
    finished = pyqtSignal()
    finished_import = pyqtSignal(list)
    show_error_message = pyqtSignal(str)
    show_message = pyqtSignal(str)
    set_status_message = pyqtSignal(str)
    finished_import_config = pyqtSignal(dict)
    finished_scan = pyqtSignal(list)
    update_progress = pyqtSignal(tuple)

    def __init__(self):
        super(Worker, self).__init__()

    @pyqtSlot()
    def generate_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["检索", "$0", "$1", "$2", "$3", "$4"])
        ws.append(["张三", "张三", "软件191", "1900000001", "一组", "1号"])
        ws.append(["李四", "李四", "软件192", "1900000002", "一组", "2号"])
        ws.append(["王五", "王五", "软件193", "1900000003", "一组", "3号"])
        ws['H1'] = "使用说明"
        ws['H2'] = "将 “检索” 列填入需要查询的文件名"
        ws['H3'] = "将列 $0, $1, $2, $3, $4 填入需要重命名的文件名格式"
        ws['H4'] = "软件将搜索目录下的所有文件，并将含有检索文件名的文件重命名"
        wb.save("exported.xlsx")
        self.finished.emit()
        win32api.ShellExecute(0, "open", "explorer.exe", "/select, exported.xlsx", "", 1)

    @pyqtSlot(str)
    def import_file(self, file_name):
        wb = load_workbook(file_name)
        ws = wb.active
        data = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
            data.append([(cell.value if cell.value else "") for cell in row])
        self.finished_import.emit(data)

    @pyqtSlot(str, dict)
    def export_config(self, file_name, config):
        with open(file_name, "w") as f:
            json.dump(config, f)
        self.show_message.emit("配置已导出")

    @pyqtSlot(str)
    def import_config(self, file_name):
        with open(file_name, "r") as f:
            config = json.load(f)
            self.finished_import_config.emit(config)

    @pyqtSlot(str, list, str)
    def scan_file(self, working_directory, name_list, rename_format):
        files = os.listdir(working_directory)
        data = []
        for row in name_list:
            name = row[0]
            rename_args = row[1:]
            for file in files:
                if name in file:
                    ext_name = os.path.splitext(file)[-1]
                    formatted_name = rename_format.format(*rename_args, extname=ext_name)
                    data.append([name, file, formatted_name])
                    break
            else:
                data.append([name, "", ""])
        self.finished_scan.emit(data)

    @pyqtSlot(dict)
    def rename_file(self, data):
        working_directory = data["working_directory"]
        name_list = data["rename_list"]
        count = 0
        for row in name_list:

            # == todo 提高速度的优化空间 ==========
            time.sleep(random.randint(0, 40) / 100)
            # ==================================

            renamed_name = (row[1], "red")
            try:
                os.rename(os.path.join(working_directory, row[1]),
                          os.path.join(working_directory, row[2]))
                status = ("完成", "blue")
                renamed_name = (row[2], "blue")
            except FileNotFoundError:
                status = ("失败：不存在", "red")
            except FileExistsError:
                status = ("失败：重复", "red")
            finally:
                count += 1
            self.set_status_message.emit("共有{}项，已完成{}项".format(len(name_list), count))
            self.update_progress.emit((row[0], renamed_name, status))
        self.show_message.emit("重命名完成")


if __name__ == '__main__':
    import sys

    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
